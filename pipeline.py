import logging
import traceback

import openpyxl
import os
import glob
from datetime import datetime
import chromedriver_autoinstaller
from selenium.common import WebDriverException
from selenium.webdriver.chrome.options import Options
import parser_it
import parser_co_jp
import parser_es
import parser_de
import parser_com
import parser_fr
import parser_co_uk
import parser_ca
import config
import uspto_search
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService  # Similar thing for firefox also!
from subprocess import CREATE_NO_WINDOW  # This flag will only be available in windows

# 日志配置
log_folder = os.path.join(os.getcwd(), 'log')
os.makedirs(log_folder, exist_ok=True)
log_file = os.path.join(log_folder, f'{datetime.now():%Y-%m-%d-%H-%M-%S}.log')
logging.basicConfig(filename=log_file, level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# 配置 ChromeDriver
# chromedriver_autoinstaller.install()
chromedriver_autoinstaller.install(cwd=True)
chrome_options = Options()

# chrome_options.add_argument('--headless')  # 无头模式,不打开浏览器窗口
chrome_options.add_argument("start-maximized")
chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
chrome_options.add_experimental_option('useAutomationExtension', False)
chrome_options.add_argument('--disable-blink-features=AutomationControlled')

# Define your own service object with the `CREATE_NO_WINDOW ` flag
# If chromedriver.exe is not in PATH, then use:
# ChromeService('/path/to/chromedriver')
chrome_service = ChromeService('chromedriver')

chrome_service.creationflags = CREATE_NO_WINDOW


# 获取 Amazon 站点 URL
def get_amazon_url(country):
    # Amazon 站点 URL
    amazon_urls = {
        '意大利': 'https://www.amazon.it/',
        '西班牙': 'https://www.amazon.es/',
        '德国': 'https://www.amazon.de/',
        '英国': 'https://www.amazon.co.uk/',
        '美国': 'https://www.amazon.com/',
        '法国': 'https://www.amazon.fr/',
        '日本': 'https://www.amazon.co.jp/',
        '加拿大': 'https://www.amazon.ca/'
    }
    return amazon_urls.get(country, '')


# 获取邮编码
def get_postal_code(country):
    # 邮编码
    postal_codes = {
        '意大利': '10129',
        '西班牙': '20570',
        '德国': '0419',
        '英国': 'BR2 9FZ',
        '美国': '10001',
        '法国': '18300',
        '日本': '116-0003',
        '加拿大': 'M2J-4A6'
    }
    return postal_codes.get(country, '')


def get_country_product_route(country):
    country_parser = {
        '意大利': parser_it.get_product_info_it,
        '西班牙': parser_es.get_product_info_es,
        '德国': parser_de.get_product_info_de,
        '英国': parser_co_uk.get_product_info_co_uk,
        '美国': parser_com.get_product_info_com,
        '法国': parser_fr.get_product_info_fr,
        '日本': parser_co_jp.get_product_info_co_jp,
        '加拿大': parser_ca.get_product_info_ca
    }

    get_product_country_info = country_parser.get(country)
    return get_product_country_info


def get_product_info(asin_list, country_name):
    # 打开 Amazon 站点
    amazon_url = get_amazon_url(country_name)
    # 输入邮编码
    postal_code = get_postal_code(country_name)

    # 随机选择 User-Agent
    user_agent = config.get_user_agents()
    chrome_options.add_argument(f'user-agent={user_agent}')

    # 创建浏览器实例
    try:
        # 假设用户没安装 Chrome, 程序在这里就停住了
        driver = webdriver.Chrome(service=chrome_service, options=chrome_options)
        # driver.maximize_window()
    except WebDriverException as e:
        logging.info(f'请确保您的设备上已经安装 Chrome 浏览器,再次重试')
        logging.error(f'出现的异常信息：{str(e)}')
        logging.error(traceback.format_exc())

    # 获取各个国家的解析器
    country_info_data = get_country_product_route(country_name)
    # 从各个国家的解析器里获取返回的数据
    info_data = country_info_data(asin_list, country_name, amazon_url, postal_code, driver)

    if driver:
        driver.quit()

    return info_data


def get_uspto_info():
    uspto_search.uspto_pipeline(chrome_service, chrome_options)
    return True


def pipeline():
    # 获取当前目录下的所有 Excel 文件
    current_dir = os.getcwd()
    # 构建 ASIN 目录的路径
    target_dir = os.path.join(current_dir, '商品ASIN')
    # 获取 ASIN 目录下的所有 Excel 文件
    excel_files = glob.glob(os.path.join(target_dir, '*.xlsx'))

    # 遍历 Excel 文件
    for excel_file in excel_files:
        # 获取国家名
        file_name = excel_file.split(".xlsx")[0]
        country_name = file_name.split("-")[-1]
        logging.info(f'开始读取->{file_name}')

        # 打开 Excel 文件
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # 获取 ASIN 列的数据
        asin_list = [cell.value for cell in ws['A']]
        # 使用列表解析去除 None 元素
        asin_list = [x for x in asin_list if x is not None]

        # 获取到的数据集合
        logging.info(f'{country_name} 开始批量执行数据采集')
        result = get_product_info(asin_list, country_name)
        logging.info(f'{country_name} 完成批量执行数据采集')
        wb.close()
        logging.info(f'结束操作->{file_name}')

        if result is None:
            logging.info(f'{country_name} amazon 反爬虫导致未能正常采集数据,请等待一段时间在重试')
            continue

        # 获取当前时间
        current_time = datetime.now()
        # 格式化为指定的时间格式
        formatted_time = current_time.strftime("%Y-%m-%d-%H-%M-%S")

        # 创建结果 Excel 文件
        result_file = f'{country_name}_符合不可售要求商品_{formatted_time}.xlsx'

        logging.info(f'开始写入->{result_file}')
        result_wb = openpyxl.Workbook()
        result_ws = result_wb.active

        # 写入结果到结果 Excel 文件
        headers = ['ASIN', '标题', '品牌名称', '评论星级', '评论数量', '大分类', '小分类', '商品链接']
        result_ws.append(headers)

        for item in result:
            row = [item.get(header) for header in headers]
            result_ws.append(row)

        # 创建文件夹路径
        folder_name = "不在售商品信息"
        folder_path = os.path.join(current_dir, folder_name)

        # 检查文件夹是否存在,如果不存在则创建
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # 保存文件到指定路径
        result_file_path = os.path.join(folder_path, result_file)
        # 保存结果 Excel 文件
        result_wb.save(result_file_path)
        result_wb.close()

        logging.info(f'已完成 {excel_file} 中的商品信息获取,结果保存在 {result_file} 中。')

    return True


if __name__ == '__main__':
    pipeline()
    # get_uspto_info()
