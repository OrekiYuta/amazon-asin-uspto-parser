import time
import traceback

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common import NoSuchElementException, WebDriverException

from selenium.webdriver.common.by import By

import config
import logging
import openpyxl
import os
import glob

from fetcher.fetcher_ca import fetch_uspto_ca
from fetcher.fetcher_com import fetch_uspto_com
from fetcher.fetcher_common import fetch_uspto_common
from fetcher.fetcher_co_jp import fetch_uspto_co_jp
from fetcher.fetcher_co_uk import fetch_uspto_co_uk

from parsers import parser_common


def get_uspto_check_urls(country):
    uspto_check_urls = {
        '意大利': 'https://www.tmdn.org/tmview/#/tmview',
        '西班牙': 'https://www.tmdn.org/tmview/#/tmview',
        '德国': 'https://www.tmdn.org/tmview/#/tmview',  # 官方 https://register.dpma.de/DPMAregister/marke/basis
        '法国': 'https://www.tmdn.org/tmview/#/tmview',

        '英国': 'https://trademarks.ipo.gov.uk/ipo-tmtext',
        '美国': 'https://tmsearch.uspto.gov/',
        '日本': 'https://www.j-platpat.inpit.go.jp/',
        '加拿大': 'https://ised-isde.canada.ca/cipo/trademark-search/srch'
    }
    return uspto_check_urls.get(country, '')


def get_uspto_checker_route(country):
    uspto_checker = {
        '意大利': fetch_uspto_common,
        '西班牙': fetch_uspto_common,
        '德国': fetch_uspto_common,
        '法国': fetch_uspto_common,

        '英国': fetch_uspto_co_uk,
        '美国': fetch_uspto_com,

        '日本': fetch_uspto_co_jp,
        '加拿大': fetch_uspto_ca
    }

    get_uspto_checker = uspto_checker.get(country)
    return get_uspto_checker


def collect_uspto(unique_brands, chrome_service, chrome_options, country_name):
    is_registered_result_dict = {}

    uspto_check_url = get_uspto_check_urls(country_name)

    logging.info(f'开始 {country_name} 品牌 ({uspto_check_url})')
    for i in range(0, len(unique_brands), 3):

        user_agent = config.get_user_agents()
        chrome_options.add_argument(f'user-agent={user_agent}')

        try:
            driver = webdriver.Chrome(service=chrome_service, options=chrome_options)
        except WebDriverException as e:
            logging.info(f'请确保您的设备上已经安装 Chrome 浏览器,再次重试')
            logging.error(f'出现的异常信息：{str(e)}')
            logging.error(traceback.format_exc())

        # 处理当前批次的品牌数据,一个实例做三次 selenium 操作就销毁
        for j in range(i, min(i + 3, len(unique_brands))):
            '''
            1.正常-执行 selenium 查询结果 -> 已注册
            2.正常-执行 selenium 查询结果 -> 未注册
            3.特殊-没有"品牌名称"- 不执行 selenium 查询结果 -> N/A
            4.特殊-有"品牌名称",但是显示的是 "Unknown"- 不执行 selenium 查询结果 -> Unknown
            '''
            current_brand = unique_brands[j]
            time.sleep(1)
            if current_brand == "Unknown":
                is_registered_value = current_brand
            elif current_brand == "未找到品牌":
                is_registered_value = "N/A"
            else:
                logging.info(f'{current_brand} 开始查询')

                # is_registered = fetch_uspto_com(current_brand, driver, uspto_check_url)
                uspto_checker_route = get_uspto_checker_route(country_name)
                is_registered = uspto_checker_route(current_brand, driver, uspto_check_url)

                if is_registered:
                    is_registered_value = "已注册"
                else:
                    is_registered_value = "未注册"

                logging.info(f'{current_brand} 查询完成 - {is_registered_value}')

            is_registered_result_dict[current_brand] = is_registered_value

        # 关闭当前批次的 driver 实例
        driver.quit()

    logging.info(f'完成 {country_name} 品牌 ({uspto_check_url})')
    return is_registered_result_dict


def fetch_uspto_com(brand, driver, uspto_check_url):
    is_registered_flag = False

    retries = 0
    max_retries = 3
    while retries < max_retries:
        try:
            driver.get(uspto_check_url)

            time.sleep(2)
            # 进入商标查询页面
            link_element = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, '/html/body/center/table[1]/tbody/tr[2]/td/font/font/a')
                )
            )
            link_element.click()

            time.sleep(2)

            # 输入商标
            search_input = driver.find_element(By.XPATH, '/html/body/form/font/table[4]/tbody/tr[1]/td/input')
            search_input.send_keys(brand)

            # 点击查询按钮
            search_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, '/html/body/form/font/table[4]/tbody/tr[4]/td/input[3]')
                )
            )

            search_button.click()
            time.sleep(2)
            break

        except Exception as e:
            parser_common.except_screenshot(driver)
            logging.error("加载查询品牌页面异常")
            logging.error(f'出现的异常信息：{str(e)}')
            logging.error(traceback.format_exc())
            time.sleep(1)
            retries += 1
            if retries > 3:
                break
            logging.info(f'查询品牌页面异常,第{retries}次重试...')

    try:

        # parser_common.fake_operation_scroll(driver)

        # 检查页面元素,查找结果
        flag_element = driver.find_element(By.XPATH, "//form[@action='/bin/showfield']")
        target_scope = flag_element.find_element(By.XPATH, './following-sibling::*[1]')

        if "Record" in target_scope.text:
            is_registered_flag = True

    except NoSuchElementException:
        logging.error(f'{brand},未找到查询结果,品牌未注册')
        logging.error(traceback.format_exc())
        pass

    return is_registered_flag


def uspto_pipeline(chrome_service, chrome_options):
    current_dir = os.getcwd()
    target_dir = os.path.join(current_dir, '不在售商品信息')
    excel_files = glob.glob(os.path.join(target_dir, '*.xlsx'))

    for excel_file in excel_files:
        file_name_with_path = excel_file.split(".xlsx")[0]
        logging.info(f'开始读取->{file_name_with_path}')
        file_name = os.path.basename(excel_file).split(".xlsx")[0]

        country_name = file_name.split("_")[0]

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # 获取指定列的数据（ASIN和品牌名称）
        asin_column = [cell.value for cell in ws['A'][1:]]  # 除开第一行标头
        brand_column = [cell.value for cell in ws['C'][1:]]  # 除开第一行标头

        # 对品牌名称进行去重
        unique_brands = list(set(brand_column))
        logging.info(f'开始批量执行品牌注册查询')
        is_registered_result_dict = collect_uspto(unique_brands, chrome_service, chrome_options, country_name)
        logging.info(f'完成批量执行品牌注册查询')

        wb.close()
        logging.info(f'结束操作->{file_name_with_path}')

        # 组合该文件的查询结果数据
        final_results = []
        for asin, brand in zip(asin_column, brand_column):
            is_registered = is_registered_result_dict.get(brand)
            final_results.append([asin, brand, is_registered])

        # 创建结果 Excel 文件
        output_folder_name = "未注册品牌信息"
        output_folder_path = os.path.join(os.getcwd(), output_folder_name)

        if not os.path.exists(output_folder_path):
            os.makedirs(output_folder_path)

        output_file_name = f'未注册品牌商品-from-{file_name}.xlsx'
        output_file_path = os.path.join(output_folder_path, output_file_name)

        logging.info(f'开始写入->{output_file_name}')
        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active

        headers = ['ASIN', '品牌名称', '是否注册品牌']
        output_sheet.append(headers)

        for result in final_results:
            output_sheet.append(result)

        output_workbook.save(output_file_path)
        output_workbook.close()
        logging.info(f'已完成 {excel_file} 中的商品注册信息查询,结果保存在 {output_file_name} 中。')


if __name__ == '__main__':
    uspto_pipeline()
